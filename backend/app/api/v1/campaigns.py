import csv
import io
import re
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.models import Campaign, CampaignExecution, CampaignLead, Customer, Organization, User
from app.schemas.schemas import CampaignLeadResponse, CampaignProgressResponse, CampaignResponse, StandardResponse
from app.services import get_voice_provider

router = APIRouter(prefix="/campaigns", tags=["Campaigns"])


def _normalize_phone(phone: Optional[str]) -> str:
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return ""
    if digits.startswith("1") and len(digits) == 11:
        return f"+{digits}"
    if digits.startswith("00"):
        return "+" + digits[2:]
    if digits.startswith("+"):
        return digits
    return f"+{digits}"


async def _get_or_create_customer(db: AsyncSession, organization_id: int, lead_name: Optional[str], phone: str, email: Optional[str]) -> Customer:
    normalized_phone = _normalize_phone(phone)
    result = await db.execute(
        select(Customer).filter(Customer.organization_id == organization_id, Customer.phone.in_([phone, normalized_phone]))
    )
    customer = result.scalars().first()
    if customer:
        return customer

    customer = Customer(
        organization_id=organization_id,
        name=lead_name or f"Lead {normalized_phone[-4:] if normalized_phone else 'Unknown'}",
        phone=normalized_phone or phone,
        email=email,
    )
    db.add(customer)
    await db.flush()
    return customer


def _parse_contacts(contents: bytes, filename: str) -> List[dict]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            if not row:
                continue
            rows.append({
                "name": (row.get("name") or row.get("full_name") or "").strip(),
                "phone": (row.get("phone") or row.get("mobile") or row.get("number") or "").strip(),
                "email": (row.get("email") or "").strip(),
            })
        return rows

    if lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        workbook = load_workbook(io.BytesIO(contents), data_only=True)
        sheet = workbook.active
        rows = []
        headers = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0:
                headers = [str(value or "").strip().lower() for value in row]
                continue
            if not any(value not in (None, "") for value in row):
                continue
            entry = {}
            for header_index, header in enumerate(headers):
                if header_index >= len(row):
                    entry[header] = ""
                else:
                    entry[header] = str(row[header_index] or "")
            rows.append({
                "name": (entry.get("name") or entry.get("full_name") or "").strip(),
                "phone": (entry.get("phone") or entry.get("mobile") or entry.get("number") or "").strip(),
                "email": (entry.get("email") or "").strip(),
            })
        return rows

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only CSV and Excel files are supported")


@router.post("/upload", response_model=StandardResponse[dict])
async def upload_campaign(
    file: UploadFile = File(...),
    campaign_name: str = Form(...),
    description: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.core.logger import logger
    logger.info("Upload endpoint entered")

    if not file.filename:
        logger.error("Upload failed: No file name provided")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Campaign upload failed: A file is required"
        )

    lower_name = (file.filename or "").lower()
    try:
        contents = await file.read()
        contacts = _parse_contacts(contents, file.filename)
    except Exception as parse_err:
        logger.error(f"Upload failed: File parsing error: {str(parse_err)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Campaign upload failed: File parsing error - {str(parse_err)}"
        )

    if not contacts:
        logger.error("Upload failed: No contacts parsed from file")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Campaign upload failed: The file contains no valid contacts"
        )

    try:
        campaign = Campaign(
            organization_id=current_user.organization_id,
            created_by_user_id=current_user.id,
            name=campaign_name,
            description=description,
            status="draft",
            source_file_name=file.filename,
            source_file_type="csv" if lower_name.endswith(".csv") else "excel",
            lead_count=len(contacts),
            payload={"uploaded_by": current_user.email},
        )
        db.add(campaign)
        await db.flush()

        for row in contacts:
            phone = row.get("phone") or ""
            if not phone:
                continue
            customer = await _get_or_create_customer(db, current_user.organization_id, row.get("name"), phone, row.get("email"))
            lead = CampaignLead(
                campaign_id=campaign.id,
                customer_id=customer.id,
                name=row.get("name"),
                phone=_normalize_phone(phone) or phone,
                email=row.get("email") or None,
                status="queued",
            )
            db.add(lead)

        await db.commit()
        await db.refresh(campaign)
    except Exception as db_err:
        logger.error(f"Upload failed: Database error: {str(db_err)}", exc_info=True)
        reason_msg = str(db_err)
        if "no such table" in reason_msg.lower():
            reason_msg = "Campaign table missing (run Alembic migrations)"
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Campaign upload failed: Database error - {reason_msg}"
        )

    logger.info(f"Upload endpoint successful: Created campaign {campaign.id} with {campaign.lead_count} leads")
    return StandardResponse(
        success=True,
        message="Campaign uploaded successfully",
        data={"campaign_id": campaign.id, "lead_count": campaign.lead_count},
    )



@router.get("", response_model=StandardResponse[list[CampaignResponse]])
async def list_campaigns(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Campaign).filter(Campaign.organization_id == current_user.organization_id).order_by(Campaign.created_at.desc())
    )
    campaigns = result.scalars().all()
    response_data = [CampaignResponse.model_validate(campaign) for campaign in campaigns]
    return StandardResponse(success=True, message="Campaigns retrieved", data=response_data)


@router.get("/{campaign_id}/progress", response_model=StandardResponse[CampaignProgressResponse])
async def get_campaign_progress(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    result = await db.execute(select(CampaignLead).filter(CampaignLead.campaign_id == campaign.id).order_by(CampaignLead.created_at.asc()))
    leads = result.scalars().all()
    campaign_payload = CampaignResponse.model_validate(campaign)
    leads_payload = [CampaignLeadResponse.model_validate(lead) for lead in leads]
    return StandardResponse(
        success=True,
        message="Campaign progress retrieved",
        data={"campaign": campaign_payload, "leads": leads_payload},
    )


@router.post("/{campaign_id}/start", response_model=StandardResponse[dict])
async def start_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status in ["running", "completed"]:
        raise HTTPException(status_code=400, detail=f"Campaign is already {campaign.status}")

    # Count pending/queued leads
    leads_res = await db.execute(
        select(CampaignLead).filter(CampaignLead.campaign_id == campaign.id, CampaignLead.status == "queued")
    )
    queued_leads = leads_res.scalars().all()
    total_queued = len(queued_leads)

    campaign.status = "running"
    campaign.started_at = campaign.started_at or __import__("datetime").datetime.utcnow()

    execution = CampaignExecution(campaign_id=campaign.id, status="running", started_at=__import__("datetime").datetime.utcnow())
    db.add(execution)
    await db.commit()

    # Enqueue to background worker
    from app.workers.campaign_worker import enqueue_campaign_job, broadcast_campaign_progress
    await enqueue_campaign_job(campaign.id)
    await broadcast_campaign_progress(campaign.id, db)

    return StandardResponse(
        success=True, 
        message="Campaign started successfully", 
        data={
            "queued": total_queued,
            "campaign_id": campaign.id,
            "status": "running"
        }
    )


@router.post("/{campaign_id}/pause", response_model=StandardResponse[dict])
async def pause_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status != "running":
        raise HTTPException(status_code=400, detail="Only running campaigns can be paused")

    campaign.status = "paused"
    
    # Update active execution status
    exec_res = await db.execute(
        select(CampaignExecution)
        .filter(CampaignExecution.campaign_id == campaign_id, CampaignExecution.status == "running")
        .order_by(CampaignExecution.created_at.desc())
        .limit(1)
    )
    execution = exec_res.scalar()
    if execution:
        execution.status = "paused"
        execution.finished_at = __import__("datetime").datetime.utcnow()

    await db.commit()

    from app.workers.campaign_worker import broadcast_campaign_progress
    await broadcast_campaign_progress(campaign.id, db)

    return StandardResponse(success=True, message="Campaign paused successfully", data={"campaign_id": campaign.id, "status": campaign.status})


@router.post("/{campaign_id}/resume", response_model=StandardResponse[dict])
async def resume_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status != "paused":
        raise HTTPException(status_code=400, detail="Only paused campaigns can be resumed")

    campaign.status = "running"
    
    execution = CampaignExecution(campaign_id=campaign.id, status="running", started_at=__import__("datetime").datetime.utcnow())
    db.add(execution)
    await db.commit()

    # Resume sequential background job
    from app.workers.campaign_worker import enqueue_campaign_job, broadcast_campaign_progress
    await enqueue_campaign_job(campaign.id)
    await broadcast_campaign_progress(campaign.id, db)

    return StandardResponse(success=True, message="Campaign resumed successfully", data={"campaign_id": campaign.id, "status": campaign.status})


@router.post("/{campaign_id}/retry-failed", response_model=StandardResponse[dict])
async def retry_failed_leads(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # Find failed leads
    leads_res = await db.execute(
        select(CampaignLead).filter(CampaignLead.campaign_id == campaign_id, CampaignLead.status == "failed")
    )
    failed_leads = leads_res.scalars().all()
    
    if not failed_leads:
        raise HTTPException(status_code=400, detail="No failed leads found to retry")

    # Reset failed leads back to queued status
    for lead in failed_leads:
        lead.status = "queued"
        lead.vapi_call_id = None
        lead.last_error = None
        
    campaign.status = "running"
    campaign.failed_count = max(0, campaign.failed_count - len(failed_leads))
    
    execution = CampaignExecution(campaign_id=campaign.id, status="running", started_at=__import__("datetime").datetime.utcnow())
    db.add(execution)
    await db.commit()

    # Enqueue background execution
    from app.workers.campaign_worker import enqueue_campaign_job, broadcast_campaign_progress
    await enqueue_campaign_job(campaign.id)
    await broadcast_campaign_progress(campaign.id, db)

    return StandardResponse(
        success=True, 
        message="Retrying failed campaign leads", 
        data={"campaign_id": campaign.id, "status": campaign.status, "queued": len(failed_leads)}
    )


@router.delete("/{campaign_id}", response_model=StandardResponse[dict])
async def delete_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.core.logger import logger
    from fastapi.responses import JSONResponse
    
    logger.info(f"Delete Flow: Endpoint Entered for campaign_id={campaign_id} by user={current_user.email}")
    try:
        campaign = await db.get(Campaign, campaign_id)
        if not campaign:
            logger.warning(f"Delete Flow: Record Not Located for campaign_id={campaign_id}")
            return JSONResponse(
                status_code=404,
                content={
                    "success": False,
                    "stage": "Database",
                    "reason": "Campaign not found"
                }
            )
            
        logger.info(f"Delete Flow: Record Located for campaign_id={campaign_id}")
        
        if campaign.organization_id != current_user.organization_id:
            logger.warning(f"Delete Flow: Permission Verification Failed for campaign_id={campaign_id} (Org mismatch: campaign={campaign.organization_id}, user={current_user.organization_id})")
            return JSONResponse(
                status_code=403,
                content={
                    "success": False,
                    "stage": "Authentication",
                    "reason": "Campaign belongs to another organization"
                }
            )
            
        logger.info(f"Delete Flow: Permission Verified for campaign_id={campaign_id}")
        
        # Cascade manual delete to prevent dependent table block issues
        logger.info("Delete Flow: Related Records Deleting...")
        leads_res = await db.execute(select(CampaignLead).filter(CampaignLead.campaign_id == campaign_id))
        leads = leads_res.scalars().all()
        for lead in leads:
            await db.delete(lead)
            
        execs_res = await db.execute(select(CampaignExecution).filter(CampaignExecution.campaign_id == campaign_id))
        execs = execs_res.scalars().all()
        for ex in execs:
            await db.delete(ex)
            
        await db.delete(campaign)
        await db.commit()
        logger.info("Delete Flow: Transaction Committed.")
        
        # Clean up Redis cache if applicable
        from app.core.config import settings
        import redis.asyncio as aioredis
        try:
            redis_url = settings.REDIS_URL or "redis://localhost:6379/0"
            r = aioredis.from_url(redis_url, socket_timeout=1.0)
            await r.delete(f"campaign_cache:{campaign_id}")
            await r.close()
            logger.info(f"Delete Flow: Redis Keys cleaned up for campaign_id={campaign_id}")
        except Exception as redis_err:
            logger.warning(f"Delete Flow: Redis clean up failed (silently bypassing): {str(redis_err)}")
            
        # Broadcast deletion event via WebSockets
        from app.voice.websocket.websocket import manager as ws_manager
        await ws_manager.broadcast_dashboard(current_user.organization_id, {
            "type": "campaign_deleted",
            "campaign_id": campaign_id
        })
        logger.info("Delete Flow: WebSocket deletion event broadcasted.")
        
        logger.info("Delete Flow: Success Response Returned.")
        return StandardResponse(success=True, message="Campaign and associated lead history successfully deleted.", data={"campaign_id": campaign_id})

    except Exception as e:
        await db.rollback()
        import traceback
        tb = traceback.format_exc()
        logger.error(f"Delete Flow failed at Database/Transaction stage in campaigns.py, delete_campaign:\n{tb}")
        return JSONResponse(
            status_code=550,  # Specific DB error status
            content={
                "success": False,
                "stage": "Database",
                "reason": f"Foreign key constraint or transactional error: {str(e)}"
            }
        )


@router.get("/{campaign_id}/leads", response_model=StandardResponse[List[CampaignLeadResponse]])
async def get_campaign_leads(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    campaign = await db.get(Campaign, campaign_id)
    if not campaign or campaign.organization_id != current_user.organization_id:
        raise HTTPException(status_code=404, detail="Campaign not found")

    leads_res = await db.execute(
        select(CampaignLead).filter(CampaignLead.campaign_id == campaign_id).order_by(CampaignLead.id.asc())
    )
    leads = leads_res.scalars().all()
    payload = [CampaignLeadResponse.model_validate(lead) for lead in leads]
    
    return StandardResponse(success=True, message="Campaign leads retrieved successfully", data=payload)
